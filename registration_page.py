from ctypes import alignment
from distutils.util import execute
from importlib.resources import contents
from msilib.schema import ComboBox
from tkinter import*
from tkinter import ttk
from tkinter.font import Font
from turtle import clear
import sqlite3
from tkinter import messagebox
from django.forms import PasswordInput
import maskpass
from pygame import font
from setuptools import Command
import re
import openpyxl

class Book:
        def __init__(self, root):
                self.root=root
                self.root.title("Inventory Management System")
                self.root.geometry("1350x700+0+0")
                self.root.config(bg="lightpink")

                lbl_title=Label(self.root, text="Book Store Inventory", relief="groove", font=("arial", 20, "bold"), fg="green", bg="silver", bd=10)
                lbl_title.pack(side=TOP, fill=X)
###########################Create variable for Database###############################
                self.store_name=StringVar()
                self.book_type=StringVar()
                self.email=StringVar()
                self.password=StringVar()
                self.C_password=StringVar()
                self.contact_no=StringVar()
                self.txt_address=StringVar()
                self.state=StringVar()
                self.city=StringVar()
                self.pin=StringVar()

                self.searchBy=StringVar()
                self.searchTxt=StringVar()
################################# featch excel data #####################################################
                self.wb=openpyxl.load_workbook("data.xlsx")
                self.s=self.wb["state_city"]
                self.row=self.s.max_row
                self.colm=self.s.max_column
                self.t=[]
                self.t2=[]
                for i in range(1,2):
                        for j in range(1, self.colm+1):
                                self.t.append(self.s.cell(i,j).value)
######################################################################################
                mng_frame=LabelFrame(self.root, bd=9, relief="ridge", bg="light pink", font=("arial", 15, "bold"), text="Register Here")
                mng_frame.place(x=5, y=53, height=620, width=575)

                m_store_name=Label(mng_frame, text="Store Name", font=("arial", 15, "bold"), bg="light green", relief="raised")
                m_store_name.place(x=10, y=20, height=40, width=230)
                txt_store_name=Entry(mng_frame, textvariable=self.store_name, font=("arial", 15), bd=3, relief="ridge", bg="light blue")
                txt_store_name.place(x=240, y=20, height=40, width=300)

                m_bookType=Label(mng_frame, text="Book Type", font=("arial", 15, "bold"), bg="light green", relief="raised")
                m_bookType.place(x=10, y=70, height=40, width=230)
                combo_bookType=ttk.Combobox(mng_frame, font=("arial", 15, "bold"), state="readonly", textvariable=self.book_type, background="light blue")
                combo_bookType["value"]=("School Books", "IT Book", "All Types")
                combo_bookType.place(x=240, y=70, height=40, width=300)


                m_email=Label(mng_frame, text="Email", font=("arial", 15, "bold"), bg="light green", relief="raised")
                m_email.place(x=10, y=120, height=40, width=230)
                txt_email=Entry(mng_frame, font=("arial", 15), textvariable=self.email, bd=3, relief="ridge", bg="light blue")
                txt_email.place(x=240, y=120, height=40, width=300)

                m_password=Label(mng_frame, text="Password", font=("arial", 15, "bold"), bg="light green", relief="raised")
                m_password.place(x=10, y=170, height=40, width=230)
                txt_password=Entry(mng_frame, textvariable=self.password, show="*", font=("arial", 15), bd=3, relief="ridge", bg="light blue")
                txt_password.place(x=240, y=170, height=40, width=300)
        
                m_confirm_pw=Label(mng_frame, text="Confirm Password", font=("arial", 15, "bold"), bg="light green", relief="raised")
                m_confirm_pw.place(x=10, y=220, height=40, width=230)
                txt_confirm_pw=Entry(mng_frame, textvariable=self.C_password, show="*", font=("arial", 15), bd=3, relief="ridge", bg="light blue")
                txt_confirm_pw.place(x=240, y=220, height=40, width=300)

                m_contact_no=Label(mng_frame, text="Contact No.", font=("arial", 15, "bold"), bg="light green", relief="raised")
                m_contact_no.place(x=10, y=270, height=40, width=230)
                txt_contact_no=Entry(mng_frame, textvariable=self.contact_no, font=("arial", 15), bd=3, relief="ridge", bg="light blue")
                txt_contact_no.place(x=240, y=270, height=40, width=300)

                m_address=Label(mng_frame, text="Address", font=("arial", 15, "bold"), bg="light green", relief="raised")
                m_address.place(x=10, y=320, height=60, width=230)
                self.txt_address=Text(mng_frame, height=3, width=37, relief="ridge", bd=3, bg="light blue")
                self.txt_address.place(x=240, y=320, height=60, width=300)
                
                self.combo_state=ttk.Combobox(mng_frame, font=("arial", 15, "bold"), textvariable=self.state, state="readonly", background="light blue")
                self.combo_state["value"]=self.t
                self.combo_state.place(x=10, y=390, height=40, width=160)
                self.combo_state.bind('<<ComboboxSelected>>', self.combofill)

                self.combo_city=ttk.Combobox(mng_frame, font=("arial", 15, "bold"), state="readonly", textvariable=self.city, background="light blue")
                self.combo_city["value"]=self.t2
                self.combo_city.place(x=175, y=390, height=40, width=160)

                m_pin_code=Label(mng_frame, text="Pin code", font=("arial", 15, "bold"), bg="light green", relief="raised")
                m_pin_code.place(x=340, y=390, height=40, width=100)
                txt_pin_code=Entry(mng_frame, textvariable=self.pin, font=("arial", 15), bd=3, relief="ridge", bg="light blue")
                txt_pin_code.place(x=440, y=390, height=40, width=100)

                b_add=Button(mng_frame, text="Submit", font=("arial", 15, "bold"), bd=5, bg="green", cursor="hand2", command=self.addBook)
                b_add.place(x=10, y=480, height=45, width="120")
                b_update=Button(mng_frame, text="Update", font=("arial", 15, "bold"), bd=5, bg="yellowgreen", cursor="hand2", command=self.updateData)
                b_update.place(x=145, y=480, height=45, width="120")
                b_delete=Button(mng_frame, text="Delete", font=("arial", 15, "bold"), bd=5, bg="red", cursor="hand2", command=self.deleteData)
                b_delete.place(x=280, y=480, height=45, width="120")
                b_clear=Button(mng_frame, text="Clear", font=("arial", 15, "bold"), bd=5, bg="blue", cursor="hand2", command=self.clear)
                b_clear.place(x=416, y=480, height=45, width="120")

########################################### showing detail data ################################################
                dtl_frame=Frame(self.root, bd=9, relief="ridge", bg="light pink")
                dtl_frame.place(x=579, y=60, height=610, width=760)

                d_lbl=Label(dtl_frame, text="Search By", font=("arial", 18, "bold"), bg="light pink")
                d_lbl.grid(row=0, column=0, padx=5, pady=10)
                combo_search=ttk.Combobox(dtl_frame, width=12, textvariable=self.searchBy, font=("arial", 15, "bold"), state="readonly", background="blue")
                combo_search["value"]=("Select", "store_name", "book_type", "city")
                combo_search.grid(row=0, column=1, padx=5, pady=10)
                txt_search=Entry(dtl_frame, font=("arial", 15, "bold"), textvariable=self.searchTxt, width=15, bd=5, relief=GROOVE)
                txt_search.grid(row=0, column=2, padx=5, pady=10)
                btn_search=Button(dtl_frame, font=("arial", 12, "bold"), width=10, text="Search", bd=5, bg="crimson", command=self.searchData)
                btn_search.grid(row=0, column=3, padx=5, pady=10)
                btn_show=Button(dtl_frame, font=("arial", 12, "bold"), width=10, text="Show All", bd=5, bg="light green", command=self.fetchData)
                btn_show.grid(row=0, column=4, padx=5, pady=10)

                tbl_frame=Frame(dtl_frame, bd=9, relief="ridge", bg="pink")
                tbl_frame.place(x=10, y=80, height=490, width=720)

                scrol_x=Scrollbar(tbl_frame, orient=HORIZONTAL)
                scrol_x.pack(side=BOTTOM, fill=X)
                scrol_y=Scrollbar(tbl_frame, orient=VERTICAL)
                scrol_y.pack(side=RIGHT, fill=Y)

                self.t_grid=ttk.Treeview(tbl_frame, columns=("store_name", "book_type", "contact_no", "address", "state", "city", "pin"), xscrollcommand=scrol_x.set, yscrollcommand=scrol_y.set)
                scrol_x.config(command=self.t_grid.xview)
                scrol_y.config(command=self.t_grid.yview)
                self.t_grid.heading("store_name",text="Store Name")
                self.t_grid.heading("book_type",text="Book Type")
                self.t_grid.heading("contact_no",text="Contact No.")
                self.t_grid.heading("address",text="Address")
                self.t_grid.heading("state", text="State")
                self.t_grid.heading("city", text="City")
                self.t_grid.heading("pin", text="Pincode")
                self.t_grid["show"]="headings"
                self.t_grid.column("store_name", width=250, anchor=CENTER)
                self.t_grid.column("book_type", width=70, anchor=CENTER)
                self.t_grid.column("contact_no", width=80, anchor=CENTER)
                self.t_grid.column("address", width=300, anchor=CENTER)
                self.t_grid.column("state", width=100, anchor=CENTER)
                self.t_grid.column("city", width=100, anchor=CENTER)
                self.t_grid.column("pin", width=50, anchor=CENTER)

                self.t_grid.pack(fill=BOTH, expand=1)
                self.t_grid.bind("<ButtonRelease-1>", self.getData)

                self.fetchData()
##############################Add Function#################################
        def combofill(self, event):
                z=self.t.index(self.combo_state.get())
                for i in range(1, self.row+1):
                        for j in range(z+1, z+2):
                                x=self.s.cell(i+1, j).value
                                if x==None:
                                        break
                                else:
                                        self.t2.append(x)
                self.combo_city.config(values=self.t2)
        def addBook(self):
                con=sqlite3.connect(database=r'bookInventory.db')
                cur=con.cursor()
                try:
                        if self.store_name.get()=="" or self.book_type.get()=="" or self.email.get()=="" or self.password.get()=="" or self.contact_no.get()=="":
                                messagebox.showerror("Error", "All fields are required!!")
                        elif self.password.get()!=self.C_password.get():
                                messagebox.showerror("Error", "Password and Confirm password is not same!!!")
                        else:
                                cur.execute("Select *from inventory where store_name=?",(self.store_name.get(),))
                                row=cur.fetchone()
                                if row!=None:
                                        messagebox.showerror("Error", "This store already exists....")
                                else:
                                        cur.execute("insert into inventory(store_name, book_type, email, password, contact_no, address, state, city, pin) values(?,?,?,?,?,?,?,?,?)", (
                                                                                        self.store_name.get(),
                                                                                        self.book_type.get(),
                                                                                        self.email.get(),
                                                                                        self.password.get(),
                                                                                        self.contact_no.get(),
                                                                                        self.txt_address.get('1.0', END),
                                                                                        self.state.get(),
                                                                                        self.city.get(),
                                                                                        self.pin.get(),
                                                                                        ))
                                        con.commit()
                                        messagebox.showinfo("Success", "Record has been insertd!!")
                                        self.fetchData()
                                        self.clear()
                                        con.close()
                except Exception as ex:
                        messagebox.showerror("Error", f"Error due to : {str(ex)}")

        def fetchData(self):
                con=sqlite3.connect(database=r"bookInventory.db")
                cur=con.cursor()
                cur.execute("select store_name, book_type, contact_no, address, state, city, pin from inventory")
                rows=cur.fetchall()
                if len(rows) != 0:
                        self.t_grid.delete(*self.t_grid.get_children())
                        for row in rows:
                                self.t_grid.insert('', END, values=row)
                        con.commit()
                con.close()
                self.clear()

        def searchData(self):
                con=sqlite3.connect(database=r"bookInventory.db")
                cur=con.cursor()
                try:
                        if self.searchBy.get()=="Select":
                                messagebox.showerror("Error", "Please choose search by option!!!", parent=self.root)
                        elif self.searchTxt.get()=="":
                                messagebox.showerror("Error", "Search area must be required!!!", parent=self.root)
                        else:
                                cur.execute("Select *from inventory where "+ self.searchBy.get()+" LIKE '%"+self.searchTxt.get()+"%'")
                                rows=cur.fetchall()
                                if len(rows)!=0:
                                        self.t_grid.delete(*self.t_grid.get_children())
                                        for row in rows:
                                                self.t_grid.insert('', END, values=row)
                                else:
                                        messagebox.showerror("Error", "Record not found!!!", parent=self.root)
                except Exception as ex:
                        messagebox.showerror("Error", f"Error due to : {str(ex)}", parent=self.root)

        def getData(self, eve):
                get_cursor_row=self.t_grid.focus()
                contents=self.t_grid.item(get_cursor_row)
                rows=contents['values']
                self.store_name.set(rows[0])
                self.book_type.set(rows[1])
                self.contact_no.set(rows[2])
                self.txt_address.delete("1.0", END)
                self.txt_address.insert(END, rows[3])
                self.state.set(rows[-3])
                self.city.set(rows[-2])
                self.pin.set(rows[-1])

        def updateData(self):
                con=sqlite3.connect(database=r"bookInventory.db")
                cur=con.cursor()
                try:
                        if self.store_name.get()=="" or self.email.get()=="" or self.password.get()=="":
                                messagebox.showerror("Error","Store name/Email/Password must be required:", parent=self.root)
                        else:
                                cur.execute("Select *from inventory where (store_name=? and email=? and password=?)", (self.store_name.get(), self.email.get(), self.password.get()))
                                row=cur.fetchone()
                                if row==None:
                                        messagebox.showerror("Error", "Invalid Store name/email/password")
                                else:
                                        cur.execute("update inventory set book_type=?, contact_no=?, address=?, state=?, city=?, pin=? where store_name=?", (
                                                                                                        self.book_type.get(),
                                                                                                        #self.email.get(),
                                                                                                        self.contact_no.get(),
                                                                                                        self.txt_address.get("1.0", END),
                                                                                                        self.state.get(),
                                                                                                        self.city.get(),
                                                                                                        self.pin.get(),
                                                                                                        self.store_name.get(),
                                                                                                        # self.email.get(),
                                                                                                        # self.password.get()
                                                                                                        ))
                                        con.commit()
                                        messagebox.showinfo("Success", "Record has been Updated!!")
                                        self.fetchData()
                                        self.clear()
                                        con.close()
                except Exception as ex:
                        messagebox.showerror("Error", f"Error due to : {str(ex)}")

        def deleteData(self):
                con=sqlite3.connect(database=r"bookInventory.db")
                cur=con.cursor()
                try:
                        if self.store_name.get()=="" or self.email.get()=="" or self.password.get()=="":
                                messagebox.showerror("Error","Store name/Email/Password must be required:", parent=self.root)
                        else:
                                cur.execute("Select *from inventory where (store_name=? and email=? and password=?)", (self.store_name.get(), self.email.get(), self.password.get()))
                                row=cur.fetchone()
                                print("row=======", row)
                                if row ==None:
                                        messagebox.showerror("Error", "This Store name/Email/Password is not valid")
                                else:
                                        cur.execute("delete from inventory where store_name=?", (self.store_name.get(),))
                                        con.commit()
                                        messagebox.showinfo("Success", "Record has been Deleted!!")
                                        con.close()
                                        self.fetchData()
                                        self.clear()
                except Exception as ex:
                        messagebox.showerror("Error", f"Error due to : {str(ex)}")

        def clear(self):
                self.store_name.set("")
                self.book_type.set("")
                self.email.set("")
                self.password.set("")
                self.C_password.set("")
                self.contact_no.set("")
                self.txt_address.delete("1.0", END)
                self.state.set("")
                self.city.set("")
                self.pin.set("")
                self.searchBy.set("")
                self.searchTxt.set("")

if __name__=="__main__":
        root=Tk()
        ob=Book(root)
        root.mainloop()

