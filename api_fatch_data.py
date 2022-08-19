from importlib.resources import contents
from logging import root
import sqlite3
#import string
from tkinter import*
from tkinter import ttk
from tkinter import messagebox
import openpyxl
import json
from wsgiref.util import application_uri
import requests
import pandas as pd

class BookTypes:
    def __init__(self, root):
        self.root=root
        self.root.title("Book Hub")
        self.root.geometry("1350x700+0+0")

################################################# variable declaration ##########################################
        self.field=StringVar()
        self.book_field=StringVar()
        self.book_id=StringVar()
        self.book_name=StringVar()
        self.book_name2=StringVar()
        self.email=StringVar()
        self.pw=StringVar()
        self.quantity=IntVar()
        # self.subject=StringVar()

################################################# featch excel data ########################################

        self.wb=openpyxl.load_workbook("data.xlsx")
        self.s=self.wb["books"]
        self.row=self.s.max_row
        self.colm=self.s.max_column
        self.t=[]
        self.t2=[]
        self.t3=[]
        for i in range(1,2):
            for j in range(1,3):
                self.t.append(self.s.cell(i,j).value)
##########################################################################################################################

        typ_frm=LabelFrame(self.root, text="Types of Books", font=("arial",25,"bold"), bd=3, bg="silver")
        typ_frm.place(x=10, y=10, height=90, width=1330)

        self.typ_combo=ttk.Combobox(typ_frm, font=("arial", 15, "bold"), state="readonly", textvariable=self.field)
        self.typ_combo["value"]=self.t
        self.typ_combo.place(x=20, y=0, height=40, width=330)
        self.typ_combo.bind('<<ComboboxSelected>>', self.combofill)

        self.typ_combo2=ttk.Combobox(typ_frm, font=("arial", 15, "bold"), state="readonly", textvariable=self.book_field)
        self.typ_combo2["value"]=self.t2
        self.typ_combo2.place(x=380, y=0, height=40, width=330)
        self.typ_combo2.bind('<<ComboboxSelected>>', self.combofill2)

        self.typ_combo3=ttk.Combobox(typ_frm, font=("arial", 15, "bold"), state="readonly", textvariable=self.book_name)
        self.typ_combo3["value"]=self.t3
        self.typ_combo3.place(x=740, y=0, height=40, width=370)

        search_btn=Button(typ_frm, text="Search", cursor="hand2", font=("arial", 15, "bold"), bd=5, bg="green", relief=GROOVE, command=self.featchAPI)
        search_btn.place(x=1120, y=0, height=42, width=150)

        save_data=LabelFrame(self.root, text="Choose Book", font=("arial",25,"bold"), bd=3, bg="silver")
        save_data.place(x=10, y=110, height=150, width=1330)

        email_lbl=Label(save_data, text="Email", font=("arial", 15, "bold"), bd=5, bg="olive")
        email_lbl.place(x=20, y=0, height=40, width=150)
        email_ent=Entry(save_data, font=("arial", 15), relief=RIDGE, textvariable=self.email)
        email_ent.place(x=170, y=0, height=40, width=350)

        pw_lbl2=Label(save_data, text="Password", font=("arial", 15, "bold"), bd=5, bg="olive")
        pw_lbl2.place(x=530, y=0, height=40, width=150)
        pw_ent2=Entry(save_data, font=("arial", 15), relief=RIDGE, textvariable=self.pw)
        pw_ent2.place(x=680, y=0, height=40, width=250)

        quantity_lbl3=Label(save_data, text="No. of Book", font=("arial", 15, "bold"), bd=5, bg="olive")
        quantity_lbl3.place(x=940, y=0, height=40, width=200)
        quantity_ent3=Entry(save_data, font=("arial", 15), relief=RIDGE, textvariable=self.quantity)
        quantity_ent3.place(x=1140, y=0, height=40, width=130)

        id_lbl=Label(save_data, text="Book ID", font=("arial", 15, "bold"), bd=5, bg="olive")
        id_lbl.place(x=20, y=50, height=40, width=150)
        self.id_ent=Entry(save_data, font=("arial", 15), relief=RIDGE, textvariable=self.book_id)
        self.id_ent.place(x=170, y=50, height=40, width=260)

        b_name_lbl=Label(save_data, text="Book Name", font=("arial", 15, "bold"), bd=5, bg="olive")
        b_name_lbl.place(x=440, y=50, height=40, width=150)
        self.b_name_ent=Entry(save_data, font=("arial", 15), relief=RIDGE, textvariable=self.book_name2)
        self.b_name_ent.place(x=590, y=50, height=40, width=520)

        save_btn=Button(save_data, text="Save", command=self.save, cursor="hand2", font=("arial", 15, "bold"), bd=5, bg="green", relief=GROOVE)
        save_btn.place(x=1120, y=50, height=42, width=150)

        book_detail=LabelFrame(self.root, text="Search Result", font=("arial",25,"bold"), bd=3, bg="silver")
        book_detail.place(x=10, y=270, height=430, width=1330)
########################################### tree view ###########################################################
        
        scrol_x=Scrollbar(book_detail, orient=HORIZONTAL)
        scrol_x.pack(side=BOTTOM, fill=X)
        scrol_y=Scrollbar(book_detail, orient=VERTICAL)
        scrol_y.pack(side=RIGHT, fill=Y)

        self.b_grid=ttk.Treeview(book_detail, columns=("book_id", "book_title"), xscrollcommand=scrol_x.set, yscrollcommand=scrol_y.set)
        scrol_x.config(command=self.b_grid.xview)
        scrol_y.config(command=self.b_grid.yview)
        self.b_grid.heading("book_id", text="ID")
        self.b_grid.heading("book_title", text="Book Name")
        self.b_grid["show"]="headings"
        self.b_grid.column("book_id", width=200, anchor=CENTER)
        self.b_grid.column("book_title", width=500, anchor=CENTER)

        self.b_grid.pack(fill=BOTH, expand=1)
        self.b_grid.bind("<ButtonRelease-1>", self.getData)

######################################### Link Combo####################
    def save(self):
        con=sqlite3.connect(database=r'bookInventory.db')
        cur=con.cursor()
        try:
            if self.email.get()=="" or self.pw.get()=="" or self.book_id.get()=="" or self.book_name2.get()=="" or self.quantity.get()=="":
                messagebox.showerror("Error", "All fields are required!!")
            else: 
                cur.execute("select * from inventory where email=? and password=?", (self.email.get(), self.pw.get()))
                row=cur.fetchone()
                print(row)
                if row == None:
                    messagebox.showerror("Error", "Email or Password is not correct!!")
                else:
                    cur.execute("Select *from bookRecordInventory where book_id=?",(self.book_id.get(),))
                    row=cur.fetchone()
                    if row!=None:
                        messagebox.showerror("Error", "This book already present in your database library!!")
                    else:
                        cur.execute("insert into bookRecordInventory(book_id, book_title, email, quantity, dept, subject) values(?,?,?,?,?,?)",(
                            self.book_id.get(),
                            self.book_name2.get(),
                            self.email.get(),
                            self.quantity.get(),
                            self.field.get(),
                            self.book_name.get(),
                        ))
                        con.commit()
                        messagebox.showinfo("Successfully", "Data inserted")
                        con.close()
        
        except Exception as ex:
            messagebox.showerror("Error", f"Error due to : {str(ex)}")

    def getData(self, eve):
        get_cursor_row=self.b_grid.focus()
        contents=self.b_grid.item(get_cursor_row)
        rows=contents['values']
        self.book_id.set(rows[-2])
        self.book_name2.set(rows[-1])
    
    def featchAPI(self):
        self.api_adr='https://www.googleapis.com/books/v1/volumes?q='
        #self.book=input("Book name : ")
        self.book=self.typ_combo3.get()
        self.url=self.api_adr+self.book
        self.json_data=requests.get(self.url).json()
        self.returnData=[]
        for jData in self.json_data['items']:
            self.viewData={
                "book_id":jData['id'],
                "book_title":jData['volumeInfo']['title'],
            }
            self.returnData.append(self.viewData)
        df=pd.DataFrame(self.returnData)
        b_data=df.values.tolist()
        self.b_grid.delete(*self.b_grid.get_children())
        for row in b_data:
            self.b_grid.insert('', END, values=row)
    def combofill(self, event):
        z=self.t.index(self.typ_combo.get())
        for i in range(1, self.row+1):
            for j in range(z+1, z+2):
                x=self.s.cell(i+1, j).value
                if x==None:
                    break
                else:
                    self.t2.append(x)
        self.typ_combo2.config(values=self.t2)

    def combofill2(self, event):
        z=self.t2.index(self.typ_combo2.get())
        for i in range(1, self.row+1):
            for j in range(z+3, z+4):
                x=self.s.cell(i+1, j).value
                if x==None:
                    break
                else:
                    self.t3.append(x)
        print("t3 chk====", self.t3)
        self.typ_combo3.config(values=self.t3)
if __name__=="__main__":  
    root=Tk() 
    ob=BookTypes(root)
    root.mainloop()
